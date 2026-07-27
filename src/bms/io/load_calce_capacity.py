"""Loader for CALCE PLN "Capacity Characterization_Initialization" exports.

These are raw Arbin cycler exports: one workbook per test-day batch, one
sheet per physical channel, one physical channel per PLN pouch cell. Two
format quirks handled here, not present in `load_calce.py`'s simpler
single-table CSV/XLSX assumption:

1. Files are saved with a `.xls` extension but are actually OOXML (xlsx)
   content — `openpyxl` refuses them based on extension alone. Worked
   around by copying to a temp path with a `.xlsx` suffix before opening,
   rather than trusting the given extension.
2. Each workbook's `Info` sheet documents the physical-channel-to-PLN-cell
   mapping only as free text in a `Comments` field (e.g.
   "Ch - 7 PLN - 1 TC - 3\nCh - 8 PLN - 2 TC - 4..."), not as a normal
   column. Parsed with a regex rather than assumed to follow a fixed
   channel-number arithmetic (there's no guarantee PLN numbering is a
   simple offset from channel numbering across all 17 files).

Confirmed (see docs/calce_dataset_note.md) that every channel in every file
has exactly one Cycle_Index. This is a single-cycle baseline
characterization test, not a multi-cycle aging dataset — do not expect
capacity-fade-over-cycles data from this loader. No temperature channel is
recorded in this export at all (`temperature_c` is returned as NaN); this
is a genuine gap in the source data, not a parsing omission.
"""

from __future__ import annotations

import re
import shutil
import tempfile
from pathlib import Path

import pandas as pd

_CHANNEL_PLN_RE = re.compile(r"ch\s*-\s*(\d+)\s*PLN\s*-\s*(\d+)", re.IGNORECASE)


def _open_workbook(path: Path):
    import openpyxl
    # Work around openpyxl's extension-based format check (see module docstring).
    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
        shutil.copyfile(path, tmp.name)
        tmp_path = tmp.name
    try:
        return openpyxl.load_workbook(tmp_path, read_only=True, data_only=True)
    finally:
        Path(tmp_path).unlink(missing_ok=True)


def _parse_channel_to_pln(info_sheet) -> dict[str, int]:
    mapping: dict[str, int] = {}
    for row in info_sheet.iter_rows(values_only=True):
        for cell in row:
            if isinstance(cell, str) and "PLN" in cell.upper():
                for chan, pln in _CHANNEL_PLN_RE.findall(cell):
                    mapping[chan.zfill(3)] = int(pln)
    return mapping


def load_calce_capacity_characterization(base_dir: str | Path) -> pd.DataFrame:
    """Load all "Capacity Characterization_Initialization" workbooks into one telemetry table.

    Returns unified-schema-shaped columns: dataset, cell_id, cycle,
    voltage_v, current_a, temperature_c (NaN — see docstring), soc,
    capacity_ah (cumulative discharge capacity within the single recorded
    cycle; NOT a per-cycle fade series).
    """
    base_dir = Path(base_dir)
    files = sorted(base_dir.glob("*.xls")) + sorted(base_dir.glob("*.xlsx"))
    if not files:
        raise ValueError(f"No .xls/.xlsx files found under {base_dir}")

    frames = []
    for fpath in files:
        wb = _open_workbook(fpath)
        if "Info" not in wb.sheetnames:
            continue
        chan_to_pln = _parse_channel_to_pln(wb["Info"])

        for sheet_name in wb.sheetnames:
            if not sheet_name.startswith("Channel_"):
                continue
            chan_num = sheet_name.split("-")[-1].strip()
            pln_id = chan_to_pln.get(chan_num)
            if pln_id is None:
                continue

            ws = wb[sheet_name]
            rows = list(ws.iter_rows(values_only=True))
            if len(rows) < 2:
                continue
            header, data_rows = rows[0], rows[1:]
            idx = {h: i for i, h in enumerate(header) if h is not None}
            required = {"Test_Time(s)", "Current(A)", "Voltage(V)", "Discharge_Capacity(Ah)", "Cycle_Index"}
            if not required.issubset(idx.keys()):
                continue

            df = pd.DataFrame(data_rows, columns=header)
            df = df.rename(columns={
                "Test_Time(s)": "time_s",
                "Current(A)": "current_a",
                "Voltage(V)": "voltage_v",
                "Discharge_Capacity(Ah)": "cumulative_discharge_ah",
                "Cycle_Index": "cycle",
            })
            df = df[["time_s", "current_a", "voltage_v", "cumulative_discharge_ah", "cycle"]].dropna(subset=["current_a", "voltage_v"])
            if df.empty:
                continue

            final_capacity = df["cumulative_discharge_ah"].max()
            df["capacity_ah"] = pd.NA
            if final_capacity and final_capacity > 0:
                df.loc[df.index[-1], "capacity_ah"] = final_capacity
                df["soc"] = (100.0 * (1 - df["cumulative_discharge_ah"] / final_capacity)).clip(0, 100)
            else:
                df["soc"] = pd.NA

            df["temperature_c"] = float("nan")  # not recorded in this export — see module docstring
            df["dataset"] = "calce_capacity_char"
            df["cell_id"] = f"CALCE_PLN_{pln_id:03d}"
            df["source_file"] = fpath.name
            frames.append(df.drop(columns=["cumulative_discharge_ah"]))

    if not frames:
        raise ValueError(f"No usable channel data parsed from {base_dir}")

    out = pd.concat(frames, ignore_index=True)
    return out.sort_values(["cell_id", "time_s"]).reset_index(drop=True)


def extract_initial_capacity_table(base_dir: str | Path) -> pd.DataFrame:
    """Convenience: one row per PLN cell with its measured initial (pre-storage) capacity."""
    telemetry = load_calce_capacity_characterization(base_dir)
    cap = telemetry.dropna(subset=["capacity_ah"])[["cell_id", "capacity_ah"]].copy()
    cap["capacity_ah"] = cap["capacity_ah"].astype(float)
    cap["pln_id"] = cap["cell_id"].str.extract(r"(\d+)$").astype(int)
    return cap.rename(columns={"capacity_ah": "initial_capacity_ah"})[["pln_id", "initial_capacity_ah"]]
